import os
import uuid
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

import config
from parsers import bank_detector, sbi_parser, hdfc_parser, scb_parser   # added scb_parser
from parsers.utils import clean_amount

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = config.UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = config.OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = config.MAX_CONTENT_LENGTH

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in config.ALLOWED_EXTENSIONS

def cleanup_old_files(folder, hours=config.CLEANUP_EXPIRY_HOURS):
    now = datetime.now().timestamp()
    for f in os.listdir(folder):
        fpath = os.path.join(folder, f)
        if os.path.isfile(fpath):
            if now - os.path.getmtime(fpath) > hours * 3600:
                os.remove(fpath)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    cleanup_old_files(app.config['UPLOAD_FOLDER'])
    cleanup_old_files(app.config['OUTPUT_FOLDER'])

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'})
    if file and allowed_file(file.filename):
        original_filename = file.filename
        filename = secure_filename(f"{uuid.uuid4()}_{original_filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        # Extract first few pages for detection
        import pdfplumber
        detection_text = ''
        with pdfplumber.open(filepath) as pdf:
            for i in range(min(3, len(pdf.pages))):
                detection_text += (pdf.pages[i].extract_text() or '') + '\n'

        bank, fmt = bank_detector.detect_bank_and_format(detection_text)

        if not bank:
            return jsonify({'success': False, 'error': 'Unsupported bank statement format. Please upload SBI, HDFC, or Standard Chartered statement.'})

        try:
            if bank == 'SBI':
                if fmt == 'corporate':
                    metadata, transactions = sbi_parser.parse_sbi_corporate(filepath)
                else:
                    metadata, transactions = sbi_parser.parse_sbi_personal(filepath)
                format_name = fmt
            elif bank == 'HDFC':
                if fmt == 'format1':
                    metadata, transactions = hdfc_parser.parse_hdfc_format1(filepath)
                else:
                    metadata, transactions = hdfc_parser.parse_hdfc_format2(filepath)
                format_name = fmt
            elif bank == 'SCB':
                metadata, transactions = scb_parser.parse_scb_standard(filepath)
                format_name = 'standard'
            else:
                return jsonify({'success': False, 'error': 'Parser not implemented'})

            # Add page count to metadata
            with pdfplumber.open(filepath) as pdf:
                metadata['page_count'] = len(pdf.pages)

            return jsonify({
                'success': True,
                'bank': bank,
                'format': format_name,
                'metadata': metadata,
                'transactions': transactions,
                'transaction_count': len(transactions),
                'original_filename': original_filename
            })

        except Exception as e:
            return jsonify({'success': False, 'error': f'Parsing failed: {str(e)}'})

    else:
        return jsonify({'success': False, 'error': 'Invalid file type. Only PDF allowed.'})

@app.route('/download_excel', methods=['POST'])
def download_excel():
    data = request.json
    bank = data['bank']
    fmt = data['format']
    metadata = data['metadata']
    transactions = data['transactions']
    original_filename = data.get('original_filename', 'statement')

    # Generate output filename
    base = os.path.splitext(original_filename)[0]
    timestamp = datetime.now().strftime('%d-%m-%Y_%H%M%S')
    output_filename = f"{base}_{timestamp}.xlsx"
    output_filename = secure_filename(output_filename)
    output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # --- Meta Data sheet ---
        meta_rows = [
            ['Bank Name', metadata.get('bank_name', bank)],
            ['Account Holder', metadata.get('account_holder', '')],
            ['Account Number', metadata.get('account_number', '')],
            ['IFSC Code', metadata.get('ifsc_code', '')],
            ['Statement Period', f"{metadata.get('from_date', '')} to {metadata.get('to_date', '')}"],
            ['Total Transactions', len(transactions)],
            ['PDF Pages', metadata.get('page_count', '')]
        ]
        meta_df = pd.DataFrame(meta_rows, columns=['Field', 'Value'])
        meta_df.to_excel(writer, sheet_name='Meta Data', index=False)

        # --- Transactions sheet ---
        if bank == 'SBI':
            columns = ['Txn Date', 'Value Date', 'Description', 'Ref No./Cheque No.', 'Debit', 'Credit', 'Balance', 'Page No']
            df = pd.DataFrame(transactions)
            df = df[['txn_date', 'value_date', 'description', 'ref_no', 'debit', 'credit', 'balance', 'page_no']]
            df.columns = columns
        elif bank == 'HDFC' and fmt == 'format1':
            columns = ['Date', 'Narration', 'Chq./Ref.No.', 'Value Dt', 'Withdrawal Amt.', 'Deposit Amt.', 'Closing Balance', 'Page No']
            df = pd.DataFrame(transactions)
            df = df[['date', 'narration', 'chq_ref', 'value_date', 'withdrawal', 'deposit', 'closing_balance', 'page_no']]
            df.columns = columns
        elif bank == 'HDFC' and fmt == 'format2':
            columns = ['Txn Date', 'Narration', 'Withdrawals', 'Deposits', 'Closing Balance', 'Page No']
            df = pd.DataFrame(transactions)
            df = df[['txn_date', 'narration', 'withdrawals', 'deposits', 'closing_balance', 'page_no']]
            df.columns = columns
        elif bank == 'SCB':
            columns = ['Date', 'Value Date', 'Description', 'Cheque', 'Deposit', 'Withdrawal', 'Balance', 'Page No']
            df = pd.DataFrame(transactions)
            df = df[['date', 'value_date', 'description', 'cheque', 'deposit', 'withdrawal', 'balance', 'page_no']]
            df.columns = columns
        else:
            df = pd.DataFrame(transactions)

        df.to_excel(writer, sheet_name='Transactions', index=False)

    # --- Apply dark blue header styling to both sheets and freeze header row in Transactions ---
    wb = load_workbook(output_path)
    
    dark_blue_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    
    # Style Meta Data sheet
    ws_meta = wb['Meta Data']
    for cell in ws_meta[1]:
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in ws_meta.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_meta.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Style Transactions sheet
    ws_trans = wb['Transactions']
    for cell in ws_trans[1]:
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_trans.freeze_panes = 'A2'
    
    for column in ws_trans.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_trans.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    wb.save(output_path)

    return jsonify({
        'success': True,
        'download_url': f'/download_file/{output_filename}',
        'filename': output_filename
    })

@app.route('/download_file/<filename>')
def download_file(filename):
    return send_file(os.path.join(app.config['OUTPUT_FOLDER'], filename),
                     as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=True)
